import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_validation_workbook():
    print("=" * 60)
    print("DAY 14: CREATING EXCEL VALIDATION WORKBOOK")
    print("=" * 60)
    
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Styles
    font_family = "Segoe UI"
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    title_font = Font(name=font_family, size=14, bold=True, color="1F497D")
    bold_font = Font(name=font_family, size=11, bold=True)
    regular_font = Font(name=font_family, size=11)
    
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    summary_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    accent_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )
    
    double_bottom_border = Border(
        top=Side(style='thin', color='1F497D'),
        bottom=Side(style='double', color='1F497D')
    )
    
    # ----------------------------------------------------
    # TAB 1: Bond Pricing Validation
    # ----------------------------------------------------
    ws1 = wb.create_sheet(title="Bond Pricing")
    ws1.views.sheetView[0].showGridLines = True
    
    ws1.append([])
    ws1.cell(row=2, column=2, value="Bond Pricing Validation (Excel vs. Python)").font = title_font
    ws1.append([])
    
    headers = [
        "Bond Name", "Settlement", "Maturity", "Coupon Rate", 
        "YTM", "Redemption", "Frequency", "Basis", 
        "Excel Price (Clean)", "Python Price (Clean)", "Difference"
    ]
    ws1.append(headers)
    
    # Test bonds data
    bonds_data = [
        ["7.26% GOI 2033", "2024-10-31", "2033-10-31", 0.0726, 0.0735, 100, 2, 4, 99.414983],
        ["91-day T-Bill", "2024-10-31", "2025-01-30", 0.0, 0.065, 100, 2, 4, 98.418952],
        ["8.5% AAA Corp 2027", "2024-10-31", "2027-10-31", 0.085, 0.082, 100, 2, 4, 100.783767]
    ]
    
    for idx, row in enumerate(bonds_data, start=5):
        ws1.cell(row=idx, column=1, value=row[0]).font = regular_font
        ws1.cell(row=idx, column=2, value=row[1]).font = regular_font
        ws1.cell(row=idx, column=3, value=row[2]).font = regular_font
        ws1.cell(row=idx, column=4, value=row[3]).number_format = '0.00%'
        ws1.cell(row=idx, column=4).font = regular_font
        ws1.cell(row=idx, column=5, value=row[4]).number_format = '0.00%'
        ws1.cell(row=idx, column=5).font = regular_font
        ws1.cell(row=idx, column=6, value=row[5]).font = regular_font
        ws1.cell(row=idx, column=7, value=row[6]).font = regular_font
        ws1.cell(row=idx, column=8, value=row[7]).font = regular_font
        
        # Excel formula for PRICE
        ws1.cell(row=idx, column=9, value=f'=PRICE(B{idx}, C{idx}, D{idx}, E{idx}, F{idx}, G{idx}, H{idx})').number_format = '0.000000'
        ws1.cell(row=idx, column=9).font = regular_font
        
        # Python value
        ws1.cell(row=idx, column=10, value=row[8]).number_format = '0.000000'
        ws1.cell(row=idx, column=10).font = regular_font
        
        # Difference formula
        ws1.cell(row=idx, column=11, value=f'=I{idx}-J{idx}').number_format = '0.000000'
        ws1.cell(row=idx, column=11).font = bold_font
        
    # Apply header styles
    for col in range(1, 12):
        cell = ws1.cell(row=4, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        
    for r in range(4, 8):
        for c in range(1, 12):
            ws1.cell(row=r, column=c).border = thin_border
            
    # ----------------------------------------------------
    # TAB 2: Duration & Convexity Validation
    # ----------------------------------------------------
    ws2 = wb.create_sheet(title="Duration & Convexity")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2.append([])
    ws2.cell(row=2, column=2, value="Duration & Convexity Validation").font = title_font
    ws2.append([])
    
    headers_dur = [
        "Bond Name", "Settlement", "Maturity", "Coupon Rate", 
        "YTM", "Frequency", "Basis", 
        "Excel Macaulay Dur", "Python Macaulay Dur", "MacDur Diff",
        "Excel Modified Dur", "Python Modified Dur", "ModDur Diff"
    ]
    ws2.append(headers_dur)
    
    # Test bonds duration data
    bonds_dur_data = [
        ["7.26% GOI 2033", "2024-10-31", "2033-10-31", 0.0726, 0.0735, 2, 4, 6.753609, 6.514212],
        ["91-day T-Bill", "2024-10-31", "2025-01-30", 0.0, 0.065, 2, 4, 0.249144, 0.241302],
        ["8.5% AAA Corp 2027", "2024-10-31", "2027-10-31", 0.085, 0.082, 2, 4, 2.711676, 2.604876]
    ]
    
    for idx, row in enumerate(bonds_dur_data, start=5):
        ws2.cell(row=idx, column=1, value=row[0]).font = regular_font
        ws2.cell(row=idx, column=2, value=row[1]).font = regular_font
        ws2.cell(row=idx, column=3, value=row[2]).font = regular_font
        ws2.cell(row=idx, column=4, value=row[3]).number_format = '0.00%'
        ws2.cell(row=idx, column=4).font = regular_font
        ws2.cell(row=idx, column=5, value=row[4]).number_format = '0.00%'
        ws2.cell(row=idx, column=5).font = regular_font
        ws2.cell(row=idx, column=6, value=row[5]).font = regular_font
        ws2.cell(row=idx, column=7, value=row[6]).font = regular_font
        
        # Macaulay Duration
        ws2.cell(row=idx, column=8, value=f'=DURATION(B{idx}, C{idx}, D{idx}, E{idx}, F{idx}, G{idx})').number_format = '0.000000'
        ws2.cell(row=idx, column=8).font = regular_font
        ws2.cell(row=idx, column=9, value=row[7]).number_format = '0.000000'
        ws2.cell(row=idx, column=9).font = regular_font
        ws2.cell(row=idx, column=10, value=f'=H{idx}-I{idx}').number_format = '0.000000'
        ws2.cell(row=idx, column=10).font = bold_font
        
        # Modified Duration
        ws2.cell(row=idx, column=11, value=f'=MDURATION(B{idx}, C{idx}, D{idx}, E{idx}, F{idx}, G{idx})').number_format = '0.000000'
        ws2.cell(row=idx, column=11).font = regular_font
        ws2.cell(row=idx, column=12, value=row[8]).number_format = '0.000000'
        ws2.cell(row=idx, column=12).font = regular_font
        ws2.cell(row=idx, column=13, value=f'=K{idx}-L{idx}').number_format = '0.000000'
        ws2.cell(row=idx, column=13).font = bold_font
        
    for col in range(1, 14):
        cell = ws2.cell(row=4, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        
    for r in range(4, 8):
        for c in range(1, 14):
            ws2.cell(row=r, column=c).border = thin_border
            
    # ----------------------------------------------------
    # TAB 3: Portfolio Aggregation
    # ----------------------------------------------------
    ws3 = wb.create_sheet(title="Portfolio Aggregation")
    ws3.views.sheetView[0].showGridLines = True
    
    ws3.cell(row=2, column=2, value="Portfolio Aggregation Validation").font = title_font
    
    # Load first 20 bonds of portfolio to avoid massive file size, or summarize
    df = pd.read_csv("bond_portfolio_krd.csv").head(30)
    
    headers_port = [
        "BondID", "ISIN", "Sector", "CreditRating", "Currency", 
        "FaceValue", "CouponRate", "Quantity", "MarketValue_INR", 
        "PortfolioWeight", "ModifiedDuration", "Convexity", 
        "Weighted ModDur", "Weighted Convexity"
    ]
    
    ws3.append([])
    ws3.append(headers_port)
    
    start_row = 4
    for idx, row in df.iterrows():
        r_idx = start_row + idx + 1
        ws3.append([
            row['BondID'], row['ISIN'], row['Sector'], row['CreditRating'], row['Currency'],
            row['FaceValue'], row['CouponRate'], row['Quantity'], row['MarketValue_INR'],
            row['PortfolioWeight'], row['ModifiedDuration'], row['Convexity']
        ])
        # Write formulas for weighted stats
        ws3.cell(row=r_idx, column=13, value=f'=K{r_idx} * J{r_idx}').number_format = '0.0000'
        ws3.cell(row=r_idx, column=14, value=f'=L{r_idx} * J{r_idx}').number_format = '0.0000'
        
    end_row = start_row + len(df)
    summary_row = end_row + 2
    
    # Add Portfolio Totals Row
    ws3.cell(row=summary_row, column=1, value="Portfolio Totals").font = bold_font
    ws3.cell(row=summary_row, column=9, value=f'=SUM(I5:I{end_row})').number_format = '#,##0.00'
    ws3.cell(row=summary_row, column=9).font = bold_font
    ws3.cell(row=summary_row, column=10, value=f'=SUM(J5:J{end_row})').number_format = '0.0000'
    ws3.cell(row=summary_row, column=10).font = bold_font
    ws3.cell(row=summary_row, column=13, value=f'=SUM(M5:M{end_row})').number_format = '0.0000'
    ws3.cell(row=summary_row, column=13).font = bold_font
    ws3.cell(row=summary_row, column=14, value=f'=SUM(N5:N{end_row})').number_format = '0.0000'
    ws3.cell(row=summary_row, column=14).font = bold_font
    
    # Format Headers and borders
    for col in range(1, 15):
        cell = ws3.cell(row=4, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        
    for r in range(4, end_row + 1):
        for c in range(1, 15):
            ws3.cell(row=r, column=c).border = thin_border
            
    for c in range(1, 15):
        ws3.cell(row=summary_row, column=c).border = double_bottom_border
        ws3.cell(row=summary_row, column=c).fill = summary_fill
        
    # ----------------------------------------------------
    # TAB 4: VaR Validation
    # ----------------------------------------------------
    ws4 = wb.create_sheet(title="VaR Validation")
    ws4.views.sheetView[0].showGridLines = True
    
    ws4.cell(row=2, column=2, value="Monte Carlo VaR vs. Parametric VaR").font = title_font
    
    # Load first 100 P&L values from scenarios for formula-driven VaR
    df_mc = pd.read_csv("monte_carlo_results.csv").head(100)
    
    # Side panel for statistics
    ws4.cell(row=4, column=2, value="Risk Model Metrics").font = bold_font
    ws4.cell(row=4, column=2).fill = summary_fill
    ws4.cell(row=4, column=3, value="Value (INR)").font = bold_font
    ws4.cell(row=4, column=3).fill = summary_fill
    
    stats_labels = [
        ["Total Portfolio MV", 32059345.97],
        ["Portfolio ModDur", 4.2031],
        ["Portfolio Convexity", 176.2455],
        ["Yield Volatility (100bps)", 0.01],
        ["95% Z-Score", 1.64485],
        ["99% Z-Score", 2.32635]
    ]
    
    for idx, row in enumerate(stats_labels, start=5):
        ws4.cell(row=idx, column=2, value=row[0]).font = regular_font
        ws4.cell(row=idx, column=3, value=row[1]).font = regular_font
        
    # Parametric VaR formulas (Duration + Convexity):
    # VaR_95 = - (Dur * Z * Vol - 0.5 * Conv * (Z*Vol)^2) * MV
    # Z_score is in C9, Vol in C8, Dur in C6, Conv in C7, MV in C5
    ws4.cell(row=11, column=2, value="95% Parametric VaR").font = bold_font
    ws4.cell(row=11, column=3, value="=-(C6 * C9 * C8 - 0.5 * C7 * POWER(C9*C8, 2)) * C5").number_format = '#,##0.00'
    ws4.cell(row=11, column=3).font = bold_font
    
    ws4.cell(row=12, column=2, value="99% Parametric VaR").font = bold_font
    ws4.cell(row=12, column=3, value="=-(C6 * C10 * C8 - 0.5 * C7 * POWER(C10*C8, 2)) * C5").number_format = '#,##0.00'
    ws4.cell(row=12, column=3).font = bold_font
    
    # MC VaR from scenario table
    ws4.cell(row=14, column=2, value="95% MC VaR (Formula)").font = bold_font
    ws4.cell(row=14, column=3, value="=PERCENTILE(G5:G104, 0.05)").number_format = '#,##0.00'
    ws4.cell(row=14, column=3).font = bold_font
    
    ws4.cell(row=15, column=2, value="95% MC CVaR (Expected Shortfall)").font = bold_font
    ws4.cell(row=15, column=3, value='=AVERAGEIF(G5:G104, "<="&C14)').number_format = '#,##0.00'
    ws4.cell(row=15, column=3).font = bold_font
    
    # Table of scenarios
    ws4.cell(row=4, column=6, value="Scenario ID").font = header_font
    ws4.cell(row=4, column=6).fill = header_fill
    ws4.cell(row=4, column=7, value="Simulated P&L (INR)").font = header_font
    ws4.cell(row=4, column=7).fill = header_fill
    
    for idx, row in df_mc.iterrows():
        r_idx = 5 + idx
        ws4.cell(row=r_idx, column=6, value=row['ScenarioID']).font = regular_font
        ws4.cell(row=r_idx, column=7, value=row['PnL_Total_INR']).font = regular_font
        ws4.cell(row=r_idx, column=7).number_format = '#,##0.00'
        
    for r in range(4, 16):
        for c in range(2, 4):
            ws4.cell(row=r, column=c).border = thin_border
            
    for r in range(4, 105):
        for c in range(6, 8):
            ws4.cell(row=r, column=c).border = thin_border
            
    # ----------------------------------------------------
    # TAB 5: Nelson-Siegel Yield Curve
    # ----------------------------------------------------
    ws5 = wb.create_sheet(title="Nelson-Siegel Fitting")
    ws5.views.sheetView[0].showGridLines = True
    
    ws5.cell(row=2, column=2, value="Nelson-Siegel Yield Curve Fit").font = title_font
    
    # Parameters
    ws5.cell(row=4, column=2, value="NS Parameter").font = bold_font
    ws5.cell(row=4, column=2).fill = summary_fill
    ws5.cell(row=4, column=3, value="Optimized Value").font = bold_font
    ws5.cell(row=4, column=3).fill = summary_fill
    
    # NS parameters from yield_curve.py:
    # beta0 = 0.080709, beta1 = -0.011183, beta2 = -0.010132, lambda = 2.000166
    ns_params = [
        ["beta0 (Level)", 0.080709],
        ["beta1 (Slope)", -0.011183],
        ["beta2 (Curvature)", -0.010132],
        ["lambda (Decay)", 2.000166]
    ]
    for idx, row in enumerate(ns_params, start=5):
        ws5.cell(row=idx, column=2, value=row[0]).font = regular_font
        ws5.cell(row=idx, column=3, value=row[1]).font = regular_font
        
    # Table of fitted yields
    ws5.cell(row=4, column=5, value="Tenor (Y)").font = header_font
    ws5.cell(row=4, column=5).fill = header_fill
    ws5.cell(row=4, column=6, value="Actual Yield").font = header_font
    ws5.cell(row=4, column=6).fill = header_fill
    ws5.cell(row=4, column=7, value="NS Fitted Yield (Formula)").font = header_font
    ws5.cell(row=4, column=7).fill = header_fill
    ws5.cell(row=4, column=8, value="Squared Error").font = header_font
    ws5.cell(row=4, column=8).fill = header_fill
    
    df_curve = pd.read_csv("yield_curve_fit_results.csv")
    for idx, row in df_curve.iterrows():
        r_idx = 5 + idx
        ws5.cell(row=r_idx, column=5, value=row['Tenor']).font = regular_font
        ws5.cell(row=r_idx, column=6, value=row['ActualYield']).font = regular_font
        ws5.cell(row=r_idx, column=6).number_format = '0.00%'
        
        # NS Formula: beta0 + beta1 * (1 - EXP(-T/lambda)) / (T/lambda) + beta2 * ((1 - EXP(-T/lambda)) / (T/lambda) - EXP(-T/lambda))
        # beta0 is C5, beta1 is C6, beta2 is C7, lambda is C8
        t_cell = f"E{r_idx}"
        formula_ns = f'=$C$5 + $C$6 * (1 - EXP(-{t_cell}/$C$8)) / ({t_cell}/$C$8) + $C$7 * ((1 - EXP(-{t_cell}/$C$8)) / ({t_cell}/$C$8) - EXP(-{t_cell}/$C$8))'
        ws5.cell(row=r_idx, column=7, value=formula_ns).font = regular_font
        ws5.cell(row=r_idx, column=7).number_format = '0.00%'
        
        # Squared error
        ws5.cell(row=r_idx, column=8, value=f'=POWER(F{r_idx}-G{r_idx}, 2)').font = regular_font
        ws5.cell(row=r_idx, column=8).number_format = '0.00000000'
        
    end_c_row = 5 + len(df_curve) - 1
    # Add RMSE row
    ws5.cell(row=end_c_row+2, column=5, value="RMSE").font = bold_font
    ws5.cell(row=end_c_row+2, column=8, value=f'=SQRT(AVERAGE(H5:H{end_c_row}))').number_format = '0.00%'
    ws5.cell(row=end_c_row+2, column=8).font = bold_font
    
    for r in range(4, 9):
        for c in range(2, 4):
            ws5.cell(row=r, column=c).border = thin_border
            
    for r in range(4, end_c_row + 1):
        for c in range(5, 9):
            ws5.cell(row=r, column=c).border = thin_border
            
    for c in range(5, 9):
        ws5.cell(row=end_c_row+2, column=c).border = double_bottom_border
        ws5.cell(row=end_c_row+2, column=c).fill = summary_fill
        
    # Auto-adjust column widths
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            # ignore formulas when measuring lengths
            for cell in col:
                val = str(cell.value or '')
                if not val.startswith('='):
                    max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    # Save file
    filename = "421_VikashKumar_ExcelValidation.xlsx"
    wb.save(filename)
    print(f"Excel Validation Workbook successfully created and saved to: {filename}")
    print("=" * 60)

if __name__ == "__main__":
    create_validation_workbook()
